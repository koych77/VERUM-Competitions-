from io import BytesIO

from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from app.models import Event, Nomination, Registration, RegistrationNomination


HEADERS = [
    "#",
    "ФИО",
    "Никнейм",
    "Дата рождения",
    "Возраст",
    "Пол",
    "Город",
    "Клуб/команда",
    "Тренер",
    "Команда",
    "Состав",
    "Телефон",
    "Тип регистрации",
    "Дата регистрации",
]

VERUM_ORANGE = "FF7900"
VERUM_DARK = "111111"
VERUM_PANEL = "1F1F1F"
VERUM_LIGHT = "F7F7F7"


REGISTRATION_TYPE_LABELS = {
    "full": "Полная",
    "short": "Короткая",
    "coach": "Ученики",
}


def _safe_sheet_name(title: str) -> str:
    cleaned = "".join(ch for ch in title if ch not in "[]:*?/\\")
    return (cleaned or "Номинация")[:31]


def build_event_export(db: Session, event: Event) -> BytesIO:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    nominations = (
        db.query(Nomination)
        .filter(Nomination.event_id == event.id)
        .order_by(Nomination.sort_order, Nomination.title)
        .all()
    )

    if not nominations:
        sheet = workbook.create_sheet("Без номинаций")
        sheet["A1"] = "В мероприятии пока нет номинаций."
        sheet["A1"].font = Font(bold=True, size=14)

    for nomination in nominations:
        sheet = workbook.create_sheet(_safe_sheet_name(nomination.title))
        sheet.append([event.title])
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
        sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor=VERUM_DARK)
        sheet["A1"].alignment = Alignment(horizontal="center")

        sheet.append([f"Номинация: {nomination.title}"])
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
        sheet["A2"].font = Font(bold=True, color=VERUM_ORANGE)
        sheet["A2"].fill = PatternFill("solid", fgColor=VERUM_PANEL)
        sheet["A2"].alignment = Alignment(horizontal="center")

        sheet.append([
            f"Дата: {event.event_date.strftime('%d.%m.%Y')}",
            f"Место: {event.place}",
            f"Возраст: {nomination.min_age}-{nomination.max_age}",
            f"Пол: {'любой' if nomination.gender_rule.value == 'any' else ('мужской' if nomination.gender_rule.value == 'male' else 'женский')}",
            f"Тип: {'командная' if getattr(nomination.battle_type, 'value', nomination.battle_type) == 'team' else 'соло'}",
        ])
        sheet.merge_cells(start_row=3, start_column=5, end_row=3, end_column=len(HEADERS))
        for cell in sheet[3]:
            cell.font = Font(color="555555")
            cell.alignment = Alignment(wrap_text=True)

        sheet.append([])
        sheet.append(HEADERS)
        header_row = 5
        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=VERUM_DARK)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color=VERUM_ORANGE))

        rows = (
            db.query(Registration)
            .join(RegistrationNomination)
            .filter(
                Registration.event_id == event.id,
                RegistrationNomination.nomination_id == nomination.id,
            )
            .order_by(Registration.full_name)
            .all()
        )
        for index, registration in enumerate(rows, start=1):
            sheet.append(
                [
                    index,
                    registration.full_name,
                    registration.nickname,
                    registration.birth_date.strftime("%d.%m.%Y"),
                    registration.age_on_event,
                    "Мужской" if registration.gender.value == "male" else "Женский",
                    registration.city or "",
                    registration.club or "",
                    registration.trainer or "",
                    registration.team_name or "",
                    registration.team_members or "",
                    registration.phone or "",
                    REGISTRATION_TYPE_LABELS.get(registration.registration_type.value, registration.registration_type.value),
                    registration.created_at.strftime("%d.%m.%Y %H:%M"),
                ]
            )
            row_index = sheet.max_row
            fill = PatternFill("solid", fgColor=VERUM_LIGHT if index % 2 == 0 else "FFFFFF")
            for cell in sheet[row_index]:
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=Side(style="hair", color="DDDDDD"))

        sheet.freeze_panes = "A6"
        sheet.auto_filter.ref = f"A5:N{max(sheet.max_row, 5)}"

        for column_index, column in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 40)
        sheet.column_dimensions["A"].width = 6
        sheet.column_dimensions["B"].width = 30
        sheet.column_dimensions["C"].width = 18
        sheet.column_dimensions["D"].width = 16
        sheet.column_dimensions["E"].width = 10
        sheet.column_dimensions["F"].width = 12
        sheet.column_dimensions["G"].width = 18
        sheet.column_dimensions["H"].width = 22
        sheet.column_dimensions["I"].width = 22
        sheet.column_dimensions["J"].width = 22
        sheet.column_dimensions["K"].width = 36
        sheet.column_dimensions["L"].width = 16
        sheet.column_dimensions["M"].width = 18
        sheet.column_dimensions["N"].width = 20

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

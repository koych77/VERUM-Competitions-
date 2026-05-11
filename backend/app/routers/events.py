from datetime import date, datetime
from io import BytesIO
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.encoders import jsonable_encoder
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session, selectinload

from app.database import get_db
from app.models import Event, EventStatus, Nomination, Registration, RegistrationNomination
from app.routers.deps import require_admin
from app.schemas import EventCreate, EventOut, EventUpdate, NominationCreate, NominationOut, NominationUpdate
from app.services.age import calculate_event_age

router = APIRouter(prefix="/api/events", tags=["events"])

ALLOWED_IMAGE_TYPES = {"image/jpeg": ".jpg", "image/png": ".png", "image/webp": ".webp"}
MAX_IMPORT_SIZE = 2 * 1024 * 1024
VERUM_ORANGE = "FF7900"
VERUM_DARK = "111111"
VERUM_PANEL = "1F1F1F"
VERUM_LIGHT = "F7F7F7"
VERUM_MUTED = "D9D9D9"


def _today() -> date:
    return date.today()


def _cell_text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _parse_date(value: object, field_name: str, errors: list[str]) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _cell_text(value)
    if not raw:
        errors.append(f"Заполните поле: {field_name}")
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    errors.append(f"Неверная дата в поле '{field_name}'. Используйте формат ДД.ММ.ГГГГ.")
    return None


def _parse_bool(value: object, default: bool = True) -> bool:
    raw = _cell_text(value).lower()
    if not raw:
        return default
    return raw in {"да", "yes", "true", "1", "y", "+"}


def _parse_status(value: object) -> EventStatus:
    raw = _cell_text(value).lower()
    return {
        "черновик": EventStatus.draft,
        "draft": EventStatus.draft,
        "открыто": EventStatus.open,
        "open": EventStatus.open,
        "закрыто": EventStatus.closed,
        "closed": EventStatus.closed,
        "архив": EventStatus.archived,
        "archived": EventStatus.archived,
    }.get(raw, EventStatus.open)


def _parse_gender_rule(value: object) -> str:
    raw = _cell_text(value).lower()
    if raw in {"мужской", "м", "male", "boys", "boy"}:
        return "male"
    if raw in {"женский", "ж", "female", "girls", "girl"}:
        return "female"
    return "any"


def _parse_int(value: object, field_name: str, errors: list[str]) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        errors.append(f"Поле '{field_name}' должно быть числом.")
        return None


def _sheet(workbook, names: list[str]):
    normalized = {sheet.title.strip().lower(): sheet for sheet in workbook.worksheets}
    for name in names:
        if name.lower() in normalized:
            return normalized[name.lower()]
    return None


@router.get("", response_model=list[EventOut])
def list_public_events(db: Session = Depends(get_db)) -> list[Event]:
    today = date.today()
    return (
        db.query(Event)
        .options(selectinload(Event.nominations))
        .filter(
            Event.status == EventStatus.open,
            Event.registration_opens_at <= today,
            Event.registration_closes_at >= today,
        )
        .order_by(Event.event_date)
        .all()
    )


@router.get("/admin", response_model=list[EventOut], dependencies=[Depends(require_admin)])
def list_admin_events(db: Session = Depends(get_db)) -> list[Event]:
    return db.query(Event).options(selectinload(Event.nominations)).order_by(Event.event_date.desc()).all()


def _style_header_row(sheet, row: int = 1) -> None:
    for cell in sheet[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=VERUM_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color=VERUM_ORANGE))


def _mark_input_cell(cell) -> None:
    cell.fill = PatternFill("solid", fgColor="FFF3E8")
    cell.border = Border(
        left=Side(style="thin", color=VERUM_ORANGE),
        right=Side(style="thin", color=VERUM_ORANGE),
        top=Side(style="thin", color=VERUM_ORANGE),
        bottom=Side(style="thin", color=VERUM_ORANGE),
    )
    cell.alignment = Alignment(vertical="center", wrap_text=True)


@router.get("/admin/import-template")
@router.get("/admin/import-template.xlsx")
def download_event_import_template() -> StreamingResponse:
    workbook = Workbook()
    event_sheet = workbook.active
    event_sheet.title = "Мероприятие"
    nominations_sheet = workbook.create_sheet("Номинации")
    instruction_sheet = workbook.create_sheet("Инструкция", 0)
    workbook.active = 1

    instruction_sheet.append(["VERUM: шаблон создания мероприятия"])
    instruction_sheet.append(["1", "На листе 'Мероприятие' заполняйте только оранжевые поля в колонке 'Ваш ответ'."])
    instruction_sheet.append(["2", "На листе 'Номинации' каждая строка — отдельная номинация. Оставьте пустые строки ниже, если они не нужны."])
    instruction_sheet.append(["3", "На телефоне нажимайте на оранжевую ячейку: для статуса, пола и да/нет появится список выбора."])
    instruction_sheet.append(["4", "Даты пишите в формате ДД.ММ.ГГГГ, например 20.09.2026."])
    instruction_sheet.append(["5", "Сохраните файл .xlsx и отправьте администратору."])
    instruction_sheet.append(["Важно", "Логотип мероприятия загружается отдельно в боте после создания мероприятия."])
    instruction_sheet.merge_cells("A1:B1")
    instruction_sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    instruction_sheet["A1"].fill = PatternFill("solid", fgColor=VERUM_DARK)
    instruction_sheet["A1"].alignment = Alignment(horizontal="center")
    instruction_sheet.column_dimensions["A"].width = 14
    instruction_sheet.column_dimensions["B"].width = 78
    instruction_sheet.sheet_view.showGridLines = False
    instruction_sheet.sheet_view.zoomScale = 125
    for row in instruction_sheet.iter_rows(min_row=2, max_row=7):
        row[0].font = Font(bold=True, color=VERUM_ORANGE)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        instruction_sheet.row_dimensions[row[0].row].height = 34

    event_rows = [
        ("Название", "VERUM CUP 2026", "Название будет видно участникам в списке мероприятий."),
        ("Дата проведения", "20.09.2026", "Формат: ДД.ММ.ГГГГ."),
        ("Место", "Минск, Дворец спорта", "Город, площадка или точный адрес."),
        ("Описание", "Открытый баттл VERUM", "Краткое описание для участников."),
        ("Дата открытия регистрации", _today().strftime("%d.%m.%Y"), "С этой даты участники увидят мероприятие."),
        ("Дата закрытия регистрации", "20.09.2026", "После этой даты регистрация закрывается."),
        ("Статус", "открыто", "Обычно используйте 'открыто'."),
        ("Чемпионат республики", "нет", "Да/нет: если да, возраст считается по году рождения, а не по точной дате рождения."),
        ("Полная регистрация", "да", "Да/нет: участник может сохранить постоянный профиль."),
        ("Короткая регистрация", "да", "Да/нет: быстрая регистрация только на это мероприятие."),
        ("Регистрация учеников", "да", "Да/нет: тренер может зарегистрировать учеников."),
    ]
    event_sheet.append(["Что заполнить", "Ваш ответ"])
    for row in event_rows:
        event_sheet.append([row[0], row[1]])
        event_sheet.cell(row=event_sheet.max_row, column=1).comment = Comment(row[2], "VERUM")
        event_sheet.cell(row=event_sheet.max_row, column=2).comment = Comment(row[2], "VERUM")
    event_sheet.column_dimensions["A"].width = 30
    event_sheet.column_dimensions["B"].width = 46
    event_sheet.freeze_panes = "A2"
    event_sheet.sheet_view.showGridLines = False
    event_sheet.sheet_view.zoomScale = 130

    nomination_headers = ["Название", "Возраст от", "Возраст до", "Пол", "Опыт", "Описание", "Активна"]
    nominations_sheet.append(nomination_headers)
    nominations_sheet.append(["Breaking Kids", 6, 9, "любой", "начинающие", "до 1 года занятий", "да"])
    nominations_sheet.append(["Breaking Junior Boys", 10, 13, "мужской", "open", "любой опыт", "да"])
    widths = [30, 13, 13, 16, 26, 44, 13]
    for index, width in enumerate(widths, start=1):
        nominations_sheet.column_dimensions[nominations_sheet.cell(row=1, column=index).column_letter].width = width

    nominations_sheet.freeze_panes = "A2"
    nominations_sheet.auto_filter.ref = "A1:G200"
    nominations_sheet.sheet_view.showGridLines = False
    nominations_sheet.sheet_view.zoomScale = 120

    header_hints = {
        "A1": "Название номинации. Пример: Breaking Kids.",
        "B1": "Минимальный возраст участника на дату мероприятия.",
        "C1": "Максимальный возраст участника на дату мероприятия.",
        "D1": "Выберите: любой, мужской или женский.",
        "E1": "Текстовое описание опыта. Пример: начинающие до 1 года.",
        "F1": "Дополнительное описание номинации.",
        "G1": "Да — номинация доступна участникам. Нет — скрыта.",
    }
    for coordinate, hint in header_hints.items():
        nominations_sheet[coordinate].comment = Comment(hint, "VERUM")

    for sheet in (event_sheet, nominations_sheet):
        _style_header_row(sheet)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=Side(style="hair", color="3A3A3A"))
        for row_index in range(2, sheet.max_row + 1):
            fill = PatternFill("solid", fgColor="FAFAFA" if row_index % 2 == 0 else "FFFFFF")
            for cell in sheet[row_index]:
                cell.fill = fill
            sheet.row_dimensions[row_index].height = 36

    for row_index in range(2, 13):
        _mark_input_cell(event_sheet.cell(row=row_index, column=2))

    for row_index in range(2, 201):
        nominations_sheet.row_dimensions[row_index].height = 34
        for column_index in range(1, 8):
            _mark_input_cell(nominations_sheet.cell(row=row_index, column=column_index))
    for row_index in range(4, 201):
        for column_index in range(1, 8):
            nominations_sheet.cell(row=row_index, column=column_index).value = None

    gender_validation = DataValidation(type="list", formula1='"любой,мужской,женский"', allow_blank=False)
    active_validation = DataValidation(type="list", formula1='"да,нет"', allow_blank=False)
    yes_no_validation = DataValidation(type="list", formula1='"да,нет"', allow_blank=False)
    status_validation = DataValidation(type="list", formula1='"открыто,черновик,закрыто,архив"', allow_blank=False)
    nominations_sheet.add_data_validation(gender_validation)
    nominations_sheet.add_data_validation(active_validation)
    event_sheet.add_data_validation(yes_no_validation)
    event_sheet.add_data_validation(status_validation)
    gender_validation.add("D2:D200")
    active_validation.add("G2:G200")
    status_validation.add("B8")
    yes_no_validation.add("B9:B12")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="verum_event_template.xlsx"'},
    )


@router.post("/admin/import-preview", dependencies=[Depends(require_admin)])
async def preview_event_import(file: UploadFile = File(...)) -> dict:
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Загрузите файл .xlsx")
    content = await file.read()
    if len(content) > MAX_IMPORT_SIZE:
        raise HTTPException(status_code=400, detail="Файл должен быть до 2 МБ")

    errors: list[str] = []
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Не удалось прочитать Excel-файл") from exc

    event_sheet = _sheet(workbook, ["Мероприятие", "Event"])
    nominations_sheet = _sheet(workbook, ["Номинации", "Nominations"])
    if event_sheet is None:
        errors.append("Не найден лист 'Мероприятие'.")
    if nominations_sheet is None:
        errors.append("Не найден лист 'Номинации'.")
    if errors:
        return {"ok": False, "errors": errors, "payload": None}

    values = {
        _cell_text(event_sheet.cell(row=row, column=1).value).lower(): event_sheet.cell(row=row, column=2).value
        for row in range(2, event_sheet.max_row + 1)
        if _cell_text(event_sheet.cell(row=row, column=1).value)
    }

    def event_value(*names: str):
        for name in names:
            if name.lower() in values:
                return values[name.lower()]
        return None

    title = _cell_text(event_value("Название", "Title"))
    place = _cell_text(event_value("Место", "Place"))
    description = _cell_text(event_value("Описание", "Description"))
    if not title:
        errors.append("Заполните название мероприятия.")
    if not place:
        errors.append("Заполните место мероприятия.")

    event_date = _parse_date(event_value("Дата проведения", "Event date"), "Дата проведения", errors)
    opens_at = _parse_date(event_value("Дата открытия регистрации", "Registration opens"), "Дата открытия регистрации", errors)
    closes_at = _parse_date(event_value("Дата закрытия регистрации", "Registration closes"), "Дата закрытия регистрации", errors)
    if opens_at and closes_at and opens_at > closes_at:
        errors.append("Дата открытия регистрации не может быть позже даты закрытия.")

    nominations = []
    for row in range(2, nominations_sheet.max_row + 1):
        nomination_title = _cell_text(nominations_sheet.cell(row=row, column=1).value)
        if not nomination_title:
            continue
        min_age = _parse_int(nominations_sheet.cell(row=row, column=2).value, f"Возраст от в строке {row}", errors)
        max_age = _parse_int(nominations_sheet.cell(row=row, column=3).value, f"Возраст до в строке {row}", errors)
        if min_age is not None and max_age is not None and min_age > max_age:
            errors.append(f"В строке {row} возраст от больше возраста до.")
        nominations.append(
            {
                "title": nomination_title,
                "min_age": min_age if min_age is not None else 0,
                "max_age": max_age if max_age is not None else 0,
                "gender_rule": _parse_gender_rule(nominations_sheet.cell(row=row, column=4).value),
                "experience": _cell_text(nominations_sheet.cell(row=row, column=5).value),
                "description": _cell_text(nominations_sheet.cell(row=row, column=6).value),
                "is_active": _parse_bool(nominations_sheet.cell(row=row, column=7).value, True),
                "sort_order": len(nominations) * 10 + 10,
            }
        )

    if not nominations:
        errors.append("Добавьте хотя бы одну номинацию.")

    payload = {
        "title": title,
        "event_date": event_date,
        "place": place,
        "description": description,
        "image_url": None,
        "registration_opens_at": opens_at,
        "registration_closes_at": closes_at,
        "status": _parse_status(event_value("Статус", "Status")),
        "is_republic_championship": _parse_bool(event_value("Чемпионат республики", "Republic championship"), False),
        "allow_full_registration": _parse_bool(event_value("Полная регистрация", "Full registration"), True),
        "allow_short_registration": _parse_bool(event_value("Короткая регистрация", "Short registration"), True),
        "allow_coach_registration": _parse_bool(event_value("Регистрация учеников", "Coach registration"), True),
        "nominations": nominations,
    }
    return {"ok": not errors, "errors": errors, "payload": jsonable_encoder(payload)}


@router.get("/{event_id}/image")
def get_event_image(event_id: int, db: Session = Depends(get_db)) -> Response:
    event = db.get(Event, event_id)
    if event is None or not event.image_content or not event.image_content_type:
        raise HTTPException(status_code=404, detail="Картинка мероприятия не найдена")
    return Response(
        content=event.image_content,
        media_type=event.image_content_type,
        headers={"Cache-Control": "public, max-age=31536000, immutable"},
    )


def get_event_with_nominations(db: Session, event_id: int) -> Event:
    event = (
        db.query(Event)
        .options(selectinload(Event.nominations))
        .filter(Event.id == event_id)
        .one_or_none()
    )
    if event is None:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")
    return event


def _normalize_event_data(payload: EventCreate | EventUpdate) -> dict:
    data = payload.model_dump(exclude={"nominations"})
    data["title"] = data["title"].strip()
    data["place"] = data["place"].strip()
    data["description"] = data.get("description", "").strip()
    return data


def _find_duplicate_event(db: Session, data: dict) -> Event | None:
    return (
        db.query(Event)
        .filter(
            Event.title == data["title"],
            Event.event_date == data["event_date"],
            Event.place == data["place"],
        )
        .order_by(Event.id.asc())
        .first()
    )


def _refresh_registration_ages(db: Session, event: Event) -> None:
    rows = db.query(Registration).filter(Registration.event_id == event.id).all()
    for registration in rows:
        registration.age_on_event = calculate_event_age(
            registration.birth_date,
            event.event_date,
            event.is_republic_championship,
        )


@router.post("/admin", response_model=EventOut, dependencies=[Depends(require_admin)])
def create_event(payload: EventCreate, db: Session = Depends(get_db)) -> Event:
    data = _normalize_event_data(payload)
    duplicate = _find_duplicate_event(db, data)
    if duplicate is not None:
        return get_event_with_nominations(db, duplicate.id)

    event = Event(**data)
    db.add(event)
    db.flush()
    for nomination_data in payload.nominations:
        db.add(Nomination(event_id=event.id, **nomination_data.model_dump()))
    db.commit()
    return get_event_with_nominations(db, event.id)


@router.put("/admin/{event_id}", response_model=EventOut, dependencies=[Depends(require_admin)])
def update_event(event_id: int, payload: EventUpdate, db: Session = Depends(get_db)) -> Event:
    event = db.get(Event, event_id)
    if event is None:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")
    for key, value in _normalize_event_data(payload).items():
        setattr(event, key, value)
    _refresh_registration_ages(db, event)
    db.commit()
    db.refresh(event)
    return event


@router.post("/admin/{event_id}/archive", response_model=EventOut, dependencies=[Depends(require_admin)])
def archive_event(event_id: int, db: Session = Depends(get_db)) -> Event:
    event = db.get(Event, event_id)
    if event is None:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")
    event.status = EventStatus.archived
    db.commit()
    db.refresh(event)
    return event


@router.delete("/admin/{event_id}", dependencies=[Depends(require_admin)])
def delete_event(event_id: int, db: Session = Depends(get_db)) -> dict[str, bool]:
    event = db.get(Event, event_id)
    if event is None:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")

    registration_ids = [row.id for row in db.query(Registration.id).filter(Registration.event_id == event_id).all()]
    if registration_ids:
        db.query(RegistrationNomination).filter(
            RegistrationNomination.registration_id.in_(registration_ids),
        ).delete(synchronize_session=False)
        db.query(Registration).filter(Registration.id.in_(registration_ids)).delete(synchronize_session=False)

    db.query(Nomination).filter(Nomination.event_id == event_id).delete(synchronize_session=False)
    db.delete(event)
    db.commit()
    return {"ok": True}


@router.post("/admin/{event_id}/image", response_model=EventOut, dependencies=[Depends(require_admin)])
async def upload_event_image(
    event_id: int,
    image: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> Event:
    event = db.get(Event, event_id)
    if event is None:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")
    if image.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(status_code=400, detail="Загрузите JPG, PNG или WEBP")

    content = await image.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Картинка должна быть до 5 МБ")

    event.image_content = content
    event.image_content_type = image.content_type
    event.image_url = f"/api/events/{event_id}/image?v={uuid4().hex}"
    db.commit()
    db.refresh(event)
    return event


@router.post("/admin/{event_id}/nominations", response_model=EventOut, dependencies=[Depends(require_admin)])
def create_nomination(event_id: int, payload: NominationCreate, db: Session = Depends(get_db)) -> Event:
    if db.get(Event, event_id) is None:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")
    if payload.min_age > payload.max_age:
        raise HTTPException(status_code=400, detail="Возраст от не может быть больше возраста до")
    nomination = Nomination(event_id=event_id, **payload.model_dump())
    db.add(nomination)
    db.commit()
    return get_event_with_nominations(db, event_id)


@router.put("/admin/nominations/{nomination_id}", response_model=EventOut, dependencies=[Depends(require_admin)])
def update_nomination(nomination_id: int, payload: NominationUpdate, db: Session = Depends(get_db)) -> Event:
    nomination = db.get(Nomination, nomination_id)
    if nomination is None:
        raise HTTPException(status_code=404, detail="Номинация не найдена")
    if payload.min_age > payload.max_age:
        raise HTTPException(status_code=400, detail="Возраст от не может быть больше возраста до")
    event_id = nomination.event_id
    for key, value in payload.model_dump().items():
        setattr(nomination, key, value)
    db.commit()
    return get_event_with_nominations(db, event_id)


@router.post("/admin/nominations/{nomination_id}/toggle", response_model=EventOut, dependencies=[Depends(require_admin)])
def toggle_nomination(nomination_id: int, db: Session = Depends(get_db)) -> Event:
    nomination = db.get(Nomination, nomination_id)
    if nomination is None:
        raise HTTPException(status_code=404, detail="Номинация не найдена")
    event_id = nomination.event_id
    nomination.is_active = not nomination.is_active
    db.commit()
    return get_event_with_nominations(db, event_id)
